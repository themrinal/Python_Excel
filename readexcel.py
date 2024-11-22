import openpyxl

# Load the Excel workbook into the 'wb' variable
wb = openpyxl.load_workbook("wrksht.xlsx")

# Print the title of the active worksheet (the sheet selected by default when the workbook is opened)
print(wb.active.title)

# Access the specific worksheet named 'First'
sheet = wb['First']

# Get the total number of rows and columns in the worksheet
rows = sheet.max_row       # Returns the number of rows with data in the worksheet
cols = sheet.max_column    # Returns the number of columns with data in the worksheet

# Uncomment the following line to see the number of rows and columns in the worksheet
# print(rows, cols)

# Loop through all rows and columns to access each cell's value
for i in range(1, rows + 1):       # Outer loop iterates through all rows (1-based indexing)
    for j in range(1, cols + 1):   # Inner loop iterates through all columns (1-based indexing)
        # Print the value of the current cell (i, j)
        print(sheet.cell(i, j).value)
